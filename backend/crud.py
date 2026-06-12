"""
SIMCUTI — Business Logic (CRUD + SAW Algorithm + Dashboard + Export)
Semua operasi database dan kalkulasi ada di sini.
"""
import io
from datetime import date, datetime, timedelta
from typing import List, Optional, Dict, Tuple
from collections import defaultdict
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, extract

from models import User, LeaveRequest, Holiday, Item
from schemas import (
    UserCreate, LeaveCreate, LeaveApproveReject,
    HolidayCreate, KaryawanSAWData, SAWResponse, ItemCreate, ItemUpdate,
    DashboardStats, MonthlyTrendItem, DepartmentStats, DashboardResponse,
    DateBreakdownItem, LeaveCalculationResponse,
)
from auth import hash_password, verify_password

# Nama hari dan bulan dalam Bahasa Indonesia
HARI_ID = {0: "Senin", 1: "Selasa", 2: "Rabu", 3: "Kamis", 4: "Jumat", 5: "Sabtu", 6: "Minggu"}
BULAN_ID = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}


# ==================== UTILS KALKULASI HARI KERJA ====================

def get_holiday_dates(db: Session) -> set:
    """Ambil semua tanggal hari libur dari database sebagai set."""
    holidays = db.query(Holiday.date).all()
    return {h.date for h in holidays}


def get_holiday_dates_with_names(db: Session) -> Dict[date, str]:
    """Ambil semua tanggal hari libur beserta namanya sebagai dict."""
    holidays = db.query(Holiday.date, Holiday.name).all()
    return {h.date: h.name for h in holidays}


def count_working_days(start_date: date, end_date: date, db: Session) -> int:
    """
    Hitung hari kerja efektif antara start_date dan end_date (inclusive).
    EXCLUDE: Sabtu, Minggu, dan hari libur nasional/cuti bersama dari DB.

    Catatan:
    - Hari libur nasional TIDAK dihitung sebagai cuti karyawan
    - Sabtu & Minggu TIDAK memotong kuota cuti
    - Pengajuan yang mencakup weekend tetap DITERIMA (hanya hari kerja yang dihitung)
    """
    if end_date < start_date:
        return 0

    holiday_dates = get_holiday_dates(db)
    working_days = 0
    current = start_date

    while current <= end_date:
        # Exclude Sabtu (weekday=5) dan Minggu (weekday=6)
        if current.weekday() < 5:
            # Exclude hari libur nasional & cuti bersama
            if current not in holiday_dates:
                working_days += 1
        current += timedelta(days=1)

    return working_days


def get_date_breakdown(
    start_date: date, end_date: date, db: Session
) -> LeaveCalculationResponse:
    """
    Hitung breakdown detail dari rentang tanggal cuti.
    Menunjukkan per-hari: hari kerja, weekend, atau libur nasional.
    """
    if end_date < start_date:
        return LeaveCalculationResponse(
            start_date=start_date, end_date=end_date,
            total_calendar_days=0, working_days=0,
            weekend_days=0, holiday_days=0, breakdown=[],
        )

    holiday_map = get_holiday_dates_with_names(db)
    breakdown = []
    working_days = 0
    weekend_days = 0
    holiday_days = 0
    current = start_date

    while current <= end_date:
        is_weekend = current.weekday() >= 5
        is_holiday = current in holiday_map
        is_working = not is_weekend and not is_holiday

        if is_working:
            working_days += 1
        elif is_weekend:
            weekend_days += 1
            # Jika weekend DAN libur nasional, tetap hitung sebagai weekend
        if is_holiday and not is_weekend:
            holiday_days += 1

        breakdown.append(DateBreakdownItem(
            date=current,
            day_name=HARI_ID[current.weekday()],
            is_working_day=is_working,
            is_weekend=is_weekend,
            is_holiday=is_holiday,
            holiday_name=holiday_map.get(current),
        ))
        current += timedelta(days=1)

    total_calendar = (end_date - start_date).days + 1
    return LeaveCalculationResponse(
        start_date=start_date,
        end_date=end_date,
        total_calendar_days=total_calendar,
        working_days=working_days,
        weekend_days=weekend_days,
        holiday_days=holiday_days,
        breakdown=breakdown,
    )


def _get_working_days_set(
    start_date: date, end_date: date, holiday_dates: set
) -> set:
    """Ambil set tanggal hari kerja dari sebuah rentang."""
    result = set()
    current = start_date
    while current <= end_date:
        if current.weekday() < 5 and current not in holiday_dates:
            result.add(current)
        current += timedelta(days=1)
    return result


def check_working_day_overlap(
    db: Session,
    user_id: int,
    start_date: date,
    end_date: date,
    exclude_leave_id: Optional[int] = None,
) -> bool:
    """
    Cek apakah ada HARI KERJA yang tumpang tindih dengan pengajuan pending lainnya.
    Berbeda dari pengecekan kalender biasa — ini hanya memeriksa hari kerja aktual,
    sehingga dua pengajuan yang overlap hanya di weekend TIDAK dianggap tumpang tindih.
    """
    holiday_dates = get_holiday_dates(db)
    new_working = _get_working_days_set(start_date, end_date, holiday_dates)
    if not new_working:
        return False

    # Ambil pengajuan pending yang secara kalender bisa overlap
    query = db.query(LeaveRequest).filter(
        LeaveRequest.user_id == user_id,
        LeaveRequest.status == "pending",
        LeaveRequest.start_date <= end_date,
        LeaveRequest.end_date >= start_date,
    )
    if exclude_leave_id:
        query = query.filter(LeaveRequest.id != exclude_leave_id)

    for req in query.all():
        existing_working = _get_working_days_set(
            req.start_date, req.end_date, holiday_dates
        )
        if new_working & existing_working:  # Ada irisan hari kerja
            return True
    return False


def get_used_leave_days(user_id: int, db: Session) -> int:
    """
    Hitung total hari cuti yang sudah digunakan (approved) oleh user pada tahun berjalan.
    Menggunakan sum() untuk efisiensi jika dalam satu tahun, 
    dan recalculate hanya jika melintasi tahun.
    """
    current_year = datetime.now().year
    year_start = date(current_year, 1, 1)
    year_end = date(current_year, 12, 31)

    leaves = db.query(LeaveRequest).filter(
        LeaveRequest.user_id == user_id,
        LeaveRequest.status == "approved",
        or_(
            and_(LeaveRequest.start_date >= year_start, LeaveRequest.start_date <= year_end),
            and_(LeaveRequest.end_date >= year_start, LeaveRequest.end_date <= year_end),
            and_(LeaveRequest.start_date < year_start, LeaveRequest.end_date > year_end)
        )
    ).all()

    total_days = 0
    for l in leaves:
        if l.start_date >= year_start and l.end_date <= year_end:
            # Full dalam tahun ini, gunakan yang tersimpan
            total_days += l.working_days
        else:
            # Melintasi batas tahun, hitung proporsional
            effective_start = max(l.start_date, year_start)
            effective_end = min(l.end_date, year_end)
            total_days += count_working_days(effective_start, effective_end, db)
        
    return total_days


# ==================== USER CRUD ====================

def create_user(db: Session, user_data: UserCreate) -> Optional[User]:
    """Buat user baru. Return None jika email sudah terdaftar."""
    existing = db.query(User).filter(User.email == user_data.email).first()
    if existing:
        return None

    user = User(
        email=user_data.email,
        name=user_data.name,
        hashed_password=hash_password(user_data.password),
        role=user_data.role or "karyawan",
        department=user_data.department,
        join_date=user_data.join_date or date.today(),
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def authenticate_user(db: Session, email: str, password: str) -> Optional[User]:
    """Verifikasi email + password. Return user jika valid, None jika tidak."""
    user = db.query(User).filter(User.email == email, User.is_active == True).first()
    if not user or not verify_password(password, user.hashed_password):
        return None
    return user


def get_all_karyawan(db: Session) -> List[User]:
    """Ambil semua user yang berperan sebagai karyawan."""
    return db.query(User).filter(User.role == "karyawan", User.is_active == True).all()


# ==================== LEAVE REQUEST CRUD ====================

def create_leave_request(
    db: Session,
    leave_data: LeaveCreate,
    user: User,
) -> Optional[LeaveRequest]:
    """
    Buat pengajuan cuti baru.
    Validasi: sisa kuota mencukupi, tidak ada pengajuan pending yang overlapping.
    """
    # Hitung hari kerja efektif (exclude Sabtu, Minggu, dan libur nasional)
    working_days = count_working_days(leave_data.start_date, leave_data.end_date, db)
    if working_days == 0:
        return None  # Semua hari adalah libur/weekend

    # Cek sisa kuota
    used = get_used_leave_days(user.id, db)
    remaining = user.annual_leave_quota - used
    if working_days > remaining:
        return None  # Kuota tidak cukup

    # Cek overlap HARI KERJA (bukan kalender) dengan pengajuan pending lainnya
    if check_working_day_overlap(db, user.id, leave_data.start_date, leave_data.end_date):
        return None  # Ada overlap hari kerja dengan pengajuan pending

    leave = LeaveRequest(
        user_id=user.id,
        start_date=leave_data.start_date,
        end_date=leave_data.end_date,
        working_days=working_days,
        reason=leave_data.reason,
        emergency_contact=leave_data.emergency_contact,
        status="pending",
    )
    db.add(leave)
    db.commit()
    db.refresh(leave)
    return leave


def get_leave_requests_by_user(
    db: Session,
    user_id: int,
    skip: int = 0,
    limit: int = 50,
    status_filter: Optional[str] = None,
) -> dict:
    """Ambil daftar pengajuan cuti milik user tertentu."""
    query = db.query(LeaveRequest).filter(LeaveRequest.user_id == user_id)
    if status_filter:
        query = query.filter(LeaveRequest.status == status_filter)
    total = query.count()
    items = query.order_by(LeaveRequest.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": items}


def get_all_leave_requests(
    db: Session,
    skip: int = 0,
    limit: int = 100,
    status_filter: Optional[str] = None,
) -> dict:
    """Ambil semua pengajuan cuti (untuk admin)."""
    query = db.query(LeaveRequest)
    if status_filter:
        query = query.filter(LeaveRequest.status == status_filter)
    total = query.count()
    items = query.order_by(LeaveRequest.created_at.desc()).offset(skip).limit(limit).all()
    return {"total": total, "items": items}


def get_leave_by_id(db: Session, leave_id: int) -> Optional[LeaveRequest]:
    """Ambil satu pengajuan cuti berdasarkan ID."""
    return db.query(LeaveRequest).filter(LeaveRequest.id == leave_id).first()


def approve_leave(
    db: Session,
    leave_id: int,
    admin: User,
) -> Optional[LeaveRequest]:
    """Admin: Approve pengajuan cuti."""
    leave = get_leave_by_id(db, leave_id)
    if not leave or leave.status != "pending":
        return None
    leave.status = "approved"
    leave.admin_id = admin.id
    db.commit()
    db.refresh(leave)
    return leave


def reject_leave(
    db: Session,
    leave_id: int,
    admin: User,
    data: LeaveApproveReject,
) -> Optional[LeaveRequest]:
    """Admin: Reject pengajuan cuti dengan catatan penolakan."""
    leave = get_leave_by_id(db, leave_id)
    if not leave or leave.status != "pending":
        return None
    leave.status = "rejected"
    leave.admin_id = admin.id
    leave.rejection_note = data.rejection_note
    db.commit()
    db.refresh(leave)
    return leave


def update_leave_request(
    db: Session,
    leave_id: int,
    leave_data,
    user: User,
) -> Optional[LeaveRequest]:
    """
    Update pengajuan cuti (hanya untuk status pending).
    User hanya bisa update pengajuan milik mereka sendiri.
    Validasi: sisa kuota mencukupi, tidak ada pengajuan pending yang overlapping.
    """
    leave = get_leave_by_id(db, leave_id)
    if not leave or leave.status != "pending" or leave.user_id != user.id:
        return None

    # Hitung hari kerja efektif untuk tanggal baru
    working_days = count_working_days(leave_data.start_date, leave_data.end_date, db)
    if working_days == 0:
        return None  # Semua hari adalah libur/weekend

    # Cek sisa kuota (dengan memperhitungkan selisih hari dengan pengajuan lama)
    used = get_used_leave_days(user.id, db)
    # Kurangi pengajuan lama dari perhitungan used days karena masih pending
    # (baru akan dikurangi kuota setelah approved)
    remaining = user.annual_leave_quota - used
    days_difference = working_days - leave.working_days
    
    if days_difference > remaining:
        return None  # Kuota tidak cukup untuk selisih hari

    # Cek overlap HARI KERJA (bukan kalender) dengan pengajuan pending lainnya
    if check_working_day_overlap(
        db, user.id, leave_data.start_date, leave_data.end_date,
        exclude_leave_id=leave_id,
    ):
        return None  # Ada overlap hari kerja dengan pengajuan pending lain

    # Update data
    leave.start_date = leave_data.start_date
    leave.end_date = leave_data.end_date
    leave.working_days = working_days
    leave.reason = leave_data.reason
    leave.emergency_contact = leave_data.emergency_contact
    
    db.commit()
    db.refresh(leave)
    return leave


def delete_leave_request(
    db: Session,
    leave_id: int,
    user: User,
) -> bool:
    """
    Hapus pengajuan cuti (hanya untuk status pending).
    User hanya bisa delete pengajuan milik mereka sendiri.
    """
    leave = get_leave_by_id(db, leave_id)
    if not leave or leave.status != "pending" or leave.user_id != user.id:
        return False
    
    db.delete(leave)
    db.commit()
    return True


# ==================== HOLIDAY CRUD ====================

def get_holidays(db: Session) -> List[Holiday]:
    """Ambil semua hari libur, diurutkan berdasarkan tanggal."""
    return db.query(Holiday).order_by(Holiday.date.asc()).all()


def create_holiday(db: Session, data: HolidayCreate) -> Optional[Holiday]:
    """Tambah hari libur baru. Return None jika tanggal sudah ada."""
    existing = db.query(Holiday).filter(Holiday.date == data.date).first()
    if existing:
        return None
    holiday = Holiday(date=data.date, name=data.name, type=data.type or "nasional")
    db.add(holiday)
    db.commit()
    db.refresh(holiday)
    return holiday


def delete_holiday(db: Session, holiday_id: int) -> bool:
    """Hapus hari libur berdasarkan ID."""
    holiday = db.query(Holiday).filter(Holiday.id == holiday_id).first()
    if not holiday:
        return False
    db.delete(holiday)
    db.commit()
    return True


# ==================== ANALYTICS & SAW ====================

def get_summary_stats(db: Session) -> dict:
    """Hitung statistik ringkasan cuti."""
    total_karyawan = db.query(User).filter(User.role == "karyawan", User.is_active == True).count()
    total_pengajuan = db.query(LeaveRequest).count()
    pending_count = db.query(LeaveRequest).filter(LeaveRequest.status == "pending").count()
    approved_count = db.query(LeaveRequest).filter(LeaveRequest.status == "approved").count()
    rejected_count = db.query(LeaveRequest).filter(LeaveRequest.status == "rejected").count()
    total_hari = db.query(func.sum(LeaveRequest.working_days)).filter(
        LeaveRequest.status == "approved"
    ).scalar() or 0

    return {
        "total_karyawan": total_karyawan,
        "total_pengajuan": total_pengajuan,
        "pending_count": pending_count,
        "approved_count": approved_count,
        "rejected_count": rejected_count,
        "total_hari_cuti_approved": total_hari,
    }


def calculate_saw(db: Session) -> SAWResponse:
    """
    Algoritma SAW (Simple Additive Weighting) untuk ranking karyawan.
    
    Kriteria & Bobot:
    - C1: Sisa Kuota Cuti (Benefit)   — 0.30
    - C2: Total Pengajuan (Cost)       — 0.20
    - C3: Jumlah Pending (Cost)        — 0.20
    - C4: % Approved (Benefit)         — 0.15
    - C5: Masa Kerja/hari (Benefit)    — 0.15
    
    Catatan: Semakin TINGGI skor SAW = karyawan lebih "disiplin" cuti
    (sisa kuota banyak, jarang pending berlebihan, approval rate tinggi, masa kerja lama)
    """
    BOBOT = {
        "sisa_kuota": 0.30,      # Benefit
        "total_pengajuan": 0.20, # Cost
        "pending": 0.20,         # Cost
        "approval_rate": 0.15,   # Benefit
        "masa_kerja": 0.15,      # Benefit
    }

    karyawans = get_all_karyawan(db)
    if not karyawans:
        return SAWResponse(bobot=BOBOT, data=[])

    today = date.today()
    raw_data = []

    for k in karyawans:
        # Hitung statistik pengajuan
        total_requests = db.query(LeaveRequest).filter(LeaveRequest.user_id == k.id).count()
        pending = db.query(LeaveRequest).filter(
            LeaveRequest.user_id == k.id, LeaveRequest.status == "pending"
        ).count()
        approved = db.query(LeaveRequest).filter(
            LeaveRequest.user_id == k.id, LeaveRequest.status == "approved"
        ).count()
        rejected = db.query(LeaveRequest).filter(
            LeaveRequest.user_id == k.id, LeaveRequest.status == "rejected"
        ).count()

        # Approval rate (dari non-pending, neutral 50% jika belum ada data)
        non_pending = total_requests - pending
        approval_rate = (approved / non_pending * 100) if non_pending > 0 else 50.0

        # Total hari cuti yang sudah digunakan
        used_days = get_used_leave_days(k.id, db)
        remaining_quota = max(k.annual_leave_quota - used_days, 0)

        # Masa kerja dalam hari
        jd = k.join_date if k.join_date else today
        masa_kerja = max((today - jd).days, 1)

        raw_data.append({
            "user": k,
            "total_requests": total_requests,
            "pending": pending,
            "approved": approved,
            "rejected": rejected,
            "approval_rate": approval_rate,
            "remaining_quota": remaining_quota,
            "used_days": used_days,
            "masa_kerja": masa_kerja,
        })

    if not raw_data:
        return SAWResponse(bobot=BOBOT, data=[])

    # ===== LANGKAH 1: Normalisasi Matriks =====
    # Benefit: nilai / max(nilai)
    # Cost: min(nilai) / nilai

    # ===== LANGKAH 1: Normalisasi Matriks (Linear Scale) =====
    max_sisa_kuota = max(r["remaining_quota"] for r in raw_data) or 1
    min_sisa_kuota = min(r["remaining_quota"] for r in raw_data)
    
    max_total_req = max(r["total_requests"] for r in raw_data) or 1
    min_total_req = min(r["total_requests"] for r in raw_data)
    
    max_pending = max(r["pending"] for r in raw_data) or 1
    min_pending = min(r["pending"] for r in raw_data)
    
    max_approval_rate = max(r["approval_rate"] for r in raw_data) or 1
    min_approval_rate = min(r["approval_rate"] for r in raw_data)
    
    max_masa_kerja = max(r["masa_kerja"] for r in raw_data) or 1
    min_masa_kerja = min(r["masa_kerja"] for r in raw_data)

    # ===== LANGKAH 2: Hitung Skor SAW =====
    result_data = []
    for i, r in enumerate(raw_data):
        k = r["user"]

        # Normalisasi Sisa Kuota (Benefit)
        n_sisa_kuota = (r["remaining_quota"] - min_sisa_kuota) / (max_sisa_kuota - min_sisa_kuota) if max_sisa_kuota > min_sisa_kuota else 1.0

        # Normalisasi Total Pengajuan (Cost)
        n_total_req = (max_total_req - r["total_requests"]) / (max_total_req - min_total_req) if max_total_req > min_total_req else 1.0

        # Normalisasi Pending (Cost)
        n_pending = (max_pending - r["pending"]) / (max_pending - min_pending) if max_pending > min_pending else 1.0

        # Normalisasi Approval Rate (Benefit)
        n_approval_rate = (r["approval_rate"] - min_approval_rate) / (max_approval_rate - min_approval_rate) if max_approval_rate > min_approval_rate else 1.0

        # Normalisasi Masa Kerja (Benefit)
        n_masa_kerja = (r["masa_kerja"] - min_masa_kerja) / (max_masa_kerja - min_masa_kerja) if max_masa_kerja > min_masa_kerja else 1.0

        # Skor SAW = Σ(bobot * nilai_normalisasi)
        saw_score = (
            BOBOT["sisa_kuota"] * n_sisa_kuota +
            BOBOT["total_pengajuan"] * n_total_req +
            BOBOT["pending"] * n_pending +
            BOBOT["approval_rate"] * n_approval_rate +
            BOBOT["masa_kerja"] * n_masa_kerja
        )

        result_data.append(KaryawanSAWData(
            user_id=k.id,
            name=k.name,
            email=k.email,
            department=k.department,
            join_date=k.join_date,
            annual_leave_quota=k.annual_leave_quota,
            total_leave_taken=r["used_days"],
            total_requests=r["total_requests"],
            pending_requests=r["pending"],
            approved_requests=r["approved"],
            rejected_requests=r["rejected"],
            approval_rate=round(r["approval_rate"], 2),
            remaining_quota=r["remaining_quota"],
            working_days=r["masa_kerja"],
            saw_score=round(saw_score, 4),
            rank=0,  # Akan diisi setelah sort
        ))

    # ===== LANGKAH 3: Rank berdasarkan skor (tertinggi = rank 1) =====
    result_data.sort(key=lambda x: x.saw_score, reverse=True)
    for i, item in enumerate(result_data):
        item.rank = i + 1

    return SAWResponse(bobot=BOBOT, data=result_data)


# ==================== ITEM CRUD ====================

def create_item(db: Session, item_data: ItemCreate, user_id: int) -> Item:
    """Create a new item."""
    db_item = Item(
        name=item_data.name,
        description=item_data.description,
        price=item_data.price,
        quantity=item_data.quantity,
        created_by=user_id
    )
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    return db_item


def get_items(db: Session, skip: int = 0, limit: int = 100, search: Optional[str] = None):
    """Get all items with optional search filter."""
    query = db.query(Item)
    if search:
        query = query.filter(Item.name.ilike(f"%{search}%"))
    total = query.count()
    items = query.offset(skip).limit(limit).all()
    return {"total": total, "items": items}


def get_item_by_id(db: Session, item_id: int) -> Optional[Item]:
    """Get item by ID."""
    return db.query(Item).filter(Item.id == item_id).first()


def update_item(db: Session, item_id: int, item_data: ItemUpdate) -> Optional[Item]:
    """Update an item."""
    db_item = get_item_by_id(db, item_id)
    if not db_item:
        return None
    
    update_data = item_data.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_item, key, value)
    
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    return db_item


def delete_item(db: Session, item_id: int) -> bool:
    """Delete an item."""
    db_item = get_item_by_id(db, item_id)
    if not db_item:
        return False
    
    db.delete(db_item)
    db.commit()
    return True


def get_items_stats(db: Session) -> dict:
    """Get statistics for all items."""
    items = db.query(Item).all()
    
    if not items:
        return {
            "total_items": 0,
            "total_value": 0.0,
            "avg_price": 0.0,
            "total_quantity": 0,
            "avg_quantity": 0.0,
        }
    
    total_items = len(items)
    total_value = sum(item.price * item.quantity for item in items)
    avg_price = sum(item.price for item in items) / total_items
    total_quantity = sum(item.quantity for item in items)
    avg_quantity = total_quantity / total_items if total_items > 0 else 0.0
    
    return {
        "total_items": total_items,
        "total_value": total_value,
        "avg_price": avg_price,
        "total_quantity": total_quantity,
        "avg_quantity": avg_quantity,
    }


# ==================== DASHBOARD ====================

def get_dashboard_data(db: Session) -> DashboardResponse:
    """
    Kumpulkan semua data untuk dashboard admin:
    - Statistik ringkasan
    - Tren bulanan (tahun berjalan)
    - Breakdown per departemen
    - Pengajuan pending terbaru
    """
    current_year = datetime.now().year

    # --- Stats ---
    total_karyawan = db.query(User).filter(
        User.role == "karyawan", User.is_active == True  # noqa: E712
    ).count()
    total_pengajuan = db.query(LeaveRequest).count()
    pending_count = db.query(LeaveRequest).filter(
        LeaveRequest.status == "pending"
    ).count()
    approved_count = db.query(LeaveRequest).filter(
        LeaveRequest.status == "approved"
    ).count()
    rejected_count = db.query(LeaveRequest).filter(
        LeaveRequest.status == "rejected"
    ).count()
    total_hari = db.query(func.sum(LeaveRequest.working_days)).filter(
        LeaveRequest.status == "approved"
    ).scalar() or 0

    avg_leave = total_hari / total_karyawan if total_karyawan > 0 else 0.0

    # Hitung rata-rata utilisasi kuota
    karyawans = db.query(User).filter(
        User.role == "karyawan", User.is_active == True  # noqa: E712
    ).all()
    total_quota = sum(k.annual_leave_quota for k in karyawans) if karyawans else 1
    quota_util = (total_hari / total_quota * 100) if total_quota > 0 else 0.0

    stats = DashboardStats(
        total_karyawan=total_karyawan,
        total_pengajuan=total_pengajuan,
        pending_count=pending_count,
        approved_count=approved_count,
        rejected_count=rejected_count,
        total_hari_cuti_approved=total_hari,
        avg_leave_per_employee=round(avg_leave, 2),
        quota_utilization_percent=round(quota_util, 2),
    )

    # --- Monthly Trends (tahun berjalan) ---
    monthly_trends = []
    for month in range(1, 13):
        month_start = date(current_year, month, 1)
        if month == 12:
            month_end = date(current_year, 12, 31)
        else:
            month_end = date(current_year, month + 1, 1) - timedelta(days=1)

        base_q = db.query(LeaveRequest).filter(
            or_(
                and_(
                    LeaveRequest.start_date >= month_start,
                    LeaveRequest.start_date <= month_end,
                ),
                and_(
                    LeaveRequest.end_date >= month_start,
                    LeaveRequest.end_date <= month_end,
                ),
            )
        )

        m_total = base_q.count()
        m_approved = base_q.filter(LeaveRequest.status == "approved").count()
        m_rejected = base_q.filter(LeaveRequest.status == "rejected").count()
        m_pending = base_q.filter(LeaveRequest.status == "pending").count()
        m_days = base_q.filter(LeaveRequest.status == "approved").with_entities(
            func.sum(LeaveRequest.working_days)
        ).scalar() or 0

        monthly_trends.append(MonthlyTrendItem(
            month=month,
            month_name=BULAN_ID[month],
            year=current_year,
            total_requests=m_total,
            approved=m_approved,
            rejected=m_rejected,
            pending=m_pending,
            total_working_days=m_days,
        ))

    # --- Department Summary ---
    dept_summary = []
    dept_groups = defaultdict(list)
    for k in karyawans:
        dept = k.department or "Tidak Ada Departemen"
        dept_groups[dept].append(k)

    for dept, members in dept_groups.items():
        member_ids = [m.id for m in members]
        d_total = db.query(LeaveRequest).filter(
            LeaveRequest.user_id.in_(member_ids)
        ).count()
        d_approved = db.query(LeaveRequest).filter(
            LeaveRequest.user_id.in_(member_ids),
            LeaveRequest.status == "approved",
        ).count()
        d_days = db.query(func.sum(LeaveRequest.working_days)).filter(
            LeaveRequest.user_id.in_(member_ids),
            LeaveRequest.status == "approved",
        ).scalar() or 0
        emp_count = len(members)

        dept_summary.append(DepartmentStats(
            department=dept,
            employee_count=emp_count,
            total_requests=d_total,
            approved_count=d_approved,
            total_days_taken=d_days,
            avg_days_per_employee=round(d_days / emp_count, 2) if emp_count > 0 else 0.0,
        ))

    # --- Recent Pending ---
    recent_pending = db.query(LeaveRequest).filter(
        LeaveRequest.status == "pending"
    ).order_by(LeaveRequest.created_at.desc()).limit(10).all()

    return DashboardResponse(
        stats=stats,
        monthly_trends=monthly_trends,
        department_summary=dept_summary,
        recent_pending=recent_pending,
    )


# ==================== EXPORT EXCEL & PDF ====================

def generate_leaves_excel(
    db: Session,
    status_filter: Optional[str] = None,
    year: Optional[int] = None,
) -> io.BytesIO:
    """
    Generate file Excel (.xlsx) berisi data pengajuan cuti.
    Mendukung filter status dan tahun.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = db.query(LeaveRequest)
    if status_filter:
        query = query.filter(LeaveRequest.status == status_filter)
    if year:
        query = query.filter(
            or_(
                extract("year", LeaveRequest.start_date) == year,
                extract("year", LeaveRequest.end_date) == year,
            )
        )
    leaves = query.order_by(LeaveRequest.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Cuti Karyawan"

    # --- Styling ---
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # --- Title Row ---
    title_year = year or datetime.now().year
    filter_label = f" (Status: {status_filter})" if status_filter else ""
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"LAPORAN DATA CUTI KARYAWAN — SIMCUTI {title_year}{filter_label}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:K2")
    ws["A2"].value = f"Digenerate pada: {datetime.now().strftime('%d %B %Y %H:%M')}"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="808080")
    ws["A2"].alignment = Alignment(horizontal="center")

    # --- Headers (Row 4) ---
    headers = [
        "No", "Nama Karyawan", "Email", "Departemen",
        "Tanggal Mulai", "Tanggal Selesai", "Hari Kerja",
        "Alasan", "Status", "Admin Pemroses", "Tanggal Pengajuan",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data Rows ---
    status_fills = {
        "approved": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "rejected": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "pending": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    }

    for idx, leave in enumerate(leaves, 1):
        row = idx + 4
        user = leave.user
        admin = leave.admin

        values = [
            idx,
            user.name if user else "-",
            user.email if user else "-",
            user.department if user else "-",
            leave.start_date.strftime("%d/%m/%Y"),
            leave.end_date.strftime("%d/%m/%Y"),
            leave.working_days,
            leave.reason,
            leave.status.upper(),
            admin.name if admin else "-",
            leave.created_at.strftime("%d/%m/%Y %H:%M") if leave.created_at else "-",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 8))
            if col == 9:  # Status column
                cell.fill = status_fills.get(leave.status, PatternFill())
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Auto-width ---
    col_widths = [5, 22, 28, 18, 14, 14, 10, 35, 12, 20, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = w

    # --- Summary Sheet ---
    ws2 = wb.create_sheet("Ringkasan")
    ws2["A1"] = "Ringkasan Data Cuti"
    ws2["A1"].font = Font(bold=True, size=14)

    summary_data = [
        ("Total Pengajuan", len(leaves)),
        ("Approved", sum(1 for l in leaves if l.status == "approved")),
        ("Rejected", sum(1 for l in leaves if l.status == "rejected")),
        ("Pending", sum(1 for l in leaves if l.status == "pending")),
        ("Total Hari Kerja (Approved)", sum(
            l.working_days for l in leaves if l.status == "approved"
        )),
    ]
    for r, (label, val) in enumerate(summary_data, 3):
        ws2.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=r, column=2, value=val)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_leaves_pdf(
    db: Session,
    status_filter: Optional[str] = None,
    year: Optional[int] = None,
) -> io.BytesIO:
    """
    Generate file PDF berisi laporan data pengajuan cuti.
    Mendukung filter status dan tahun.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    query = db.query(LeaveRequest)
    if status_filter:
        query = query.filter(LeaveRequest.status == status_filter)
    if year:
        query = query.filter(
            or_(
                extract("year", LeaveRequest.start_date) == year,
                extract("year", LeaveRequest.end_date) == year,
            )
        )
    leaves = query.order_by(LeaveRequest.created_at.desc()).all()

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle", parent=styles["Title"],
        fontSize=16, spaceAfter=6, alignment=TA_CENTER,
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle", parent=styles["Normal"],
        fontSize=9, textColor=colors.grey, alignment=TA_CENTER,
        spaceAfter=12,
    )
    cell_style = ParagraphStyle(
        "CellStyle", parent=styles["Normal"], fontSize=7, leading=9,
    )

    elements = []

    # --- Title ---
    title_year = year or datetime.now().year
    filter_label = f" — Status: {status_filter.upper()}" if status_filter else ""
    elements.append(Paragraph(
        f"LAPORAN DATA CUTI KARYAWAN — SIMCUTI {title_year}{filter_label}",
        title_style,
    ))
    elements.append(Paragraph(
        f"Digenerate pada: {datetime.now().strftime('%d %B %Y %H:%M')}",
        subtitle_style,
    ))

    # --- Summary ---
    approved_count = sum(1 for l in leaves if l.status == "approved")
    rejected_count = sum(1 for l in leaves if l.status == "rejected")
    pending_count = sum(1 for l in leaves if l.status == "pending")
    total_days = sum(l.working_days for l in leaves if l.status == "approved")

    summary_data = [
        ["Total Pengajuan", "Approved", "Rejected", "Pending", "Total Hari Cuti"],
        [str(len(leaves)), str(approved_count), str(rejected_count),
         str(pending_count), str(total_days)],
    ]
    summary_table = Table(summary_data, colWidths=[120, 80, 80, 80, 100])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#D9E2F3")),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 12))

    # --- Data Table ---
    table_headers = [
        "No", "Nama", "Departemen", "Tgl Mulai", "Tgl Selesai",
        "Hari Kerja", "Alasan", "Status", "Admin", "Tgl Ajuan",
    ]
    data_rows = [table_headers]

    for idx, leave in enumerate(leaves, 1):
        user = leave.user
        admin = leave.admin
        reason_text = Paragraph(
            (leave.reason or "-")[:80], cell_style
        )
        data_rows.append([
            str(idx),
            user.name if user else "-",
            user.department if user else "-",
            leave.start_date.strftime("%d/%m/%Y"),
            leave.end_date.strftime("%d/%m/%Y"),
            str(leave.working_days),
            reason_text,
            leave.status.upper(),
            admin.name if admin else "-",
            leave.created_at.strftime("%d/%m/%Y") if leave.created_at else "-",
        ])

    col_widths = [25, 85, 70, 55, 55, 35, 170, 50, 80, 55]
    data_table = Table(data_rows, colWidths=col_widths, repeatRows=1)

    # Status-based coloring
    status_colors = {
        "approved": colors.HexColor("#C6EFCE"),
        "rejected": colors.HexColor("#FFC7CE"),
        "pending": colors.HexColor("#FFEB9C"),
    }
    table_style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (6, 1), (6, -1), "LEFT"),  # Alasan left-aligned
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]
    for row_idx, leave in enumerate(leaves, 1):
        bg = status_colors.get(leave.status)
        if bg:
            table_style_cmds.append(("BACKGROUND", (7, row_idx), (7, row_idx), bg))

    data_table.setStyle(TableStyle(table_style_cmds))
    elements.append(data_table)

    doc.build(elements)
    output.seek(0)
    return output